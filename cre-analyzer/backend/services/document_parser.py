"""
Document parsing service using Claude API.
Supports PDF (Offering Memorandums, T12 financials) and Excel (T12, rent rolls).
"""

import base64
import json
import os
import io
from typing import Dict, Any, Optional
import anthropic
import openpyxl


client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
MODEL = "claude-sonnet-4-6"
MAX_TOKENS = 16384


OM_PROMPT = """You are a commercial real estate analyst extracting data from an Offering Memorandum (OM).

Extract all available information and return a JSON object with this structure:
{
  "property_info": {
    "name": "",
    "address": "",
    "asset_type": "",
    "units": 0,
    "sqft": 0,
    "year_built": 0,
    "purchase_price": 0,
    "asking_price": 0,
    "market": "",
    "submarket": "",
    "sponsor_projected_noi": 0
  },
  "unit_mix": [
    {"type": "1BR/1BA", "count": 0, "sqft": 0, "market_rent": 0}
  ],
  "market_summary": "",
  "sponsor_highlights": [],
  "flags": []
}

For missing or ambiguous fields, include in "flags" array with a note. Return only valid JSON.
"""

T12_PROMPT = """You are a commercial real estate analyst extracting trailing 12-month (T12) financial data.

Extract all income and expense line items and return a JSON object with this structure:
{
  "t12_data": {
    "gross_potential_rent": 0,
    "vacancy_loss": 0,
    "concessions": 0,
    "bad_debt": 0,
    "other_income": 0,
    "effective_gross_income": 0,
    "property_taxes": 0,
    "insurance": 0,
    "management_fee": 0,
    "maintenance_repairs": 0,
    "utilities": 0,
    "payroll": 0,
    "general_admin": 0,
    "marketing": 0,
    "capex_reserves": 0,
    "other_expenses": 0,
    "total_expenses": 0,
    "noi": 0
  },
  "period": "TTM ending MM/YYYY",
  "notes": "",
  "one_time_items": [],
  "flags": []
}

Normalize management fee to a reasonable market rate if it appears abnormal. Flag one-time items.
All values should be ANNUAL totals in dollars. Return only valid JSON.
"""

RENT_ROLL_PROMPT = """You are a commercial real estate analyst extracting rent roll data.

Extract all unit information and return a JSON object with this structure:
{
  "rent_roll": [
    {
      "unit_number": "",
      "unit_type": "1BR/1BA",
      "sqft": 0,
      "market_rent": 0,
      "current_rent": 0,
      "lease_start": "YYYY-MM-DD or null",
      "lease_end": "YYYY-MM-DD or null",
      "status": "Occupied"
    }
  ],
  "summary": {
    "total_units": 0,
    "occupied_units": 0,
    "vacant_units": 0,
    "mtm_units": 0,
    "occupancy_pct": 0,
    "avg_market_rent": 0,
    "avg_current_rent": 0,
    "total_market_gpr": 0,
    "total_current_gpr": 0
  },
  "flags": []
}

Status should be one of: Occupied, Vacant, MTM (month-to-month).
All rents should be MONTHLY per unit. Return only valid JSON.
"""


def _excel_to_text(file_bytes: bytes) -> str:
    """Convert Excel to structured text for Claude."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"\n=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                row_text = "\t".join(str(c) if c is not None else "" for c in row)
                lines.append(row_text)
    return "\n".join(lines)


def _parse_response(response) -> Dict[str, Any]:
    """Extract JSON from a Claude response, surfacing truncation clearly."""
    if response.stop_reason == "max_tokens":
        raise ValueError(
            "Document is too large to extract in a single pass (model output was "
            f"truncated at {MAX_TOKENS} tokens). Try splitting the document or "
            "uploading a summary rent roll."
        )
    text = response.content[0].text.strip()
    # Strip markdown code fences if present
    if text.startswith("```"):
        text = text.split("```")[1]
        if text.startswith("json"):
            text = text[4:]
    return json.loads(text)


def _call_claude_pdf(pdf_bytes: bytes, prompt: str) -> Dict[str, Any]:
    """Send PDF to Claude for extraction."""
    b64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")
    response = client.messages.create(
        model=MODEL,
        max_tokens=MAX_TOKENS,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "document",
                        "source": {
                            "type": "base64",
                            "media_type": "application/pdf",
                            "data": b64,
                        },
                    },
                    {"type": "text", "text": prompt},
                ],
            }
        ],
    )
    return _parse_response(response)


def _call_claude_text(content: str, prompt: str) -> Dict[str, Any]:
    """Send text content to Claude for extraction."""
    response = client.messages.create(
        model=MODEL,
        max_tokens=MAX_TOKENS,
        messages=[
            {
                "role": "user",
                "content": f"{prompt}\n\nDocument content:\n{content}",
            }
        ],
    )
    return _parse_response(response)


def parse_offering_memorandum(file_bytes: bytes, filename: str) -> Dict[str, Any]:
    """Parse an Offering Memorandum PDF."""
    return _call_claude_pdf(file_bytes, OM_PROMPT)


def parse_t12(file_bytes: bytes, filename: str) -> Dict[str, Any]:
    """Parse T12 financials from PDF or Excel."""
    if filename.lower().endswith((".xlsx", ".xls", ".csv")):
        text = _excel_to_text(file_bytes)
        return _call_claude_text(text, T12_PROMPT)
    return _call_claude_pdf(file_bytes, T12_PROMPT)


def parse_rent_roll(file_bytes: bytes, filename: str) -> Dict[str, Any]:
    """Parse rent roll from PDF or Excel."""
    if filename.lower().endswith((".xlsx", ".xls", ".csv")):
        text = _excel_to_text(file_bytes)
        return _call_claude_text(text, RENT_ROLL_PROMPT)
    return _call_claude_pdf(file_bytes, RENT_ROLL_PROMPT)


def get_demo_data() -> Dict[str, Any]:
    """Return pre-populated demo deal data (100-unit multifamily)."""
    return {
        "property_info": {
            "name": "Sunset Ridge Apartments",
            "address": "4500 Sunset Ridge Dr, Austin, TX 78741",
            "asset_type": "Multifamily",
            "units": 100,
            "sqft": 85000,
            "year_built": 1998,
            "purchase_price": 15_000_000,
            "asking_price": 15_500_000,
            "market": "Austin, TX",
            "submarket": "South Austin",
            "sponsor_projected_noi": 825_000,
        },
        "t12_data": {
            "gross_potential_rent": 1_260_000,
            "vacancy_loss": 63_000,
            "concessions": 12_600,
            "bad_debt": 6_300,
            "other_income": 60_000,
            "effective_gross_income": 1_238_100,
            "property_taxes": 138_000,
            "insurance": 42_000,
            "management_fee": 49_524,
            "maintenance_repairs": 65_000,
            "utilities": 38_000,
            "payroll": 78_000,
            "general_admin": 28_000,
            "marketing": 15_000,
            "capex_reserves": 25_000,
            "other_expenses": 7_576,
            "total_expenses": 486_100,
            "noi": 752_000,
        },
        "rent_roll": [
            {
                "unit_number": f"{i+101}",
                "unit_type": "1BR/1BA" if i % 3 != 2 else "2BR/2BA",
                "sqft": 750 if i % 3 != 2 else 1050,
                "market_rent": 1050 if i % 3 != 2 else 1400,
                "current_rent": 1020 if i % 3 != 2 else 1350,
                "lease_start": "2024-01-01",
                "lease_end": "2024-12-31",
                "status": "Occupied" if i < 95 else "Vacant",
            }
            for i in range(100)
        ],
    }
